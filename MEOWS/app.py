import pathlib
from tkinter import messagebox
from typing import Tuple
import customtkinter as ctk
import tkinter as tk

import openpyxl

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearance()
        self.todo_sistema()

    def layout_config(self):
        self.title("MEOWS - Hospital Regional do Sertão Central")    
        self.geometry("700x500")

    def appearance(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema:", bg_color="transparent", text_color=['#000', "#fff"])
        self.lb_apm.place(x=50, y=430)
        
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_appearance)
        self.opt_apm.place(x=50, y=460)

    def todo_sistema(self):

        frame = ctk.CTkFrame(self,width=700,height=100,corner_radius=0,bg_color="teal",fg_color="teal")
        frame.place(x=0,y=10) 

        title = ctk.CTkLabel(frame,text="MEOWS",text_color="#fff",font=("Century Gothic bold",55),fg_color="teal")
        title.place(x=240,y=28)

        span = ctk.CTkLabel(self,text="Por favor, preencha os dados: ",text_color=["#000","#fff"],font=("Century Gothic bold",16))
        span.place(x=50,y=130)

        planilha = pathlib.Path("Gestantes.xlsx")
        if(planilha.exists()):
                pass
        else:
                planilha=openpyxl.Workbook()
                folha = planilha.active
                folha["A1"]="Nome completo"
                folha["B1"]="Data de Nascimento"
                folha["C1"]="Pressão Diástolica"
                folha["D1"]="Pressão Respiratória"
                folha["E1"]="Pressão Sistólica"
                folha["F1"]="Frequência Respiratória"
                folha["G1"]="Temperatura"
                folha["H1"]="Nível de Consciência"
                folha["I1"]="Saturação"
                folha["J1"]="Data Atual"
                folha["K1"]="N.Prontuário"
               
                planilha.save("Gestantes.xlsx") 

        def submit():

            
            # pegando os dados dos entrys
            nome = nome_value.get()
            idade = idade_value.get()
            diast = diast_value.get()
            resp = respira_value.get()
            sist = sist_value.get()
            freq = freq_value.get()
            temperatura = temp_value.get()
            nv = nv_box.get() 
            satura = sat_value.get()
            data_atual = dat_value.get()
            prontuario = pront_value.get()

            if(nome=="" or idade =="" or diast=="" or resp=="" or sist=="" or freq=="" or temperatura=="" or nv=="" or satura=="" or data_atual=="" or prontuario==""):
                 messagebox.showerror("MEOWS","ERRO!\nPor favor, insira todos os dados!")
            else:   

                planilha = openpyxl.load_workbook("Gestantes.xlsx")
                folha = planilha.active
                folha.cell(column = 1, row = folha.max_row+1,value= nome)
                folha.cell(column = 2, row = folha.max_row,value= idade)
                folha.cell(column = 3, row = folha.max_row,value= diast)
                folha.cell(column = 4, row = folha.max_row,value= resp)
                folha.cell(column = 5, row = folha.max_row,value= sist)
                folha.cell(column = 6, row = folha.max_row,value= freq)
                folha.cell(column = 7, row = folha.max_row,value= temperatura)
                folha.cell(column = 8, row = folha.max_row,value= nv)
                folha.cell(column = 9, row = folha.max_row,value= satura)
                folha.cell(column = 10, row = folha.max_row,value= data_atual)
                folha.cell(column = 11, row = folha.max_row,value= prontuario)

                planilha.save(r"Gestantes.xlsx")
                messagebox.showinfo("Sistema","Dados salvos com sucesso!")


        def clear():
            nome_value.set("")
            idade_value.set("")
            diast_value.set("")
            respira_value.set("")
            sist_value.set("")
            freq_value.set("")
            temp_value.set("")
            sat_value.set("")
            dat_value.set("")
            pront_value.set("")

        def calc():
            pass     

        # Text Variables
        nome_value = tk.StringVar()
        idade_value = tk.StringVar()
        diast_value = tk.StringVar()
        respira_value = tk.StringVar()
        sist_value = tk.StringVar()
        freq_value = tk.StringVar()
        temp_value = tk.StringVar()
        sat_value = tk.StringVar()
        dat_value = tk.StringVar()
        pront_value = tk.StringVar()

        #entrys

        nome_entry = ctk.CTkEntry(self,width=330,font=("Century Gothic bold",16),textvariable=nome_value,fg_color="transparent")
        nome_entry.place(x=100,y=190)

        idade_entry = ctk.CTkEntry(self,width=100,font=("Century Gothic bold",16),textvariable=idade_value,fg_color="transparent")
        idade_entry.place(x=590,y=190)

        diastolica_entry = ctk.CTkEntry(self,width=100,font=("Century Gothic bold",16),textvariable=diast_value,fg_color="transparent")
        diastolica_entry.place(x=205,y=240)

        respiratoria_entry = ctk.CTkEntry(self,width=100,font=("Century Gothic bold",16),textvariable=respira_value,fg_color="transparent")
        respiratoria_entry.place(x=205,y=280)

        sistolica_entry = ctk.CTkEntry(self,width=100,font=("Century Gothic bold",16),textvariable=sist_value,fg_color="transparent")
        sistolica_entry.place(x=205,y=320)

        freq_entry = ctk.CTkEntry(self,width=100,font=("Century Gothic bold",16),textvariable=freq_value,fg_color="transparent")
        freq_entry.place(x=205,y=360)

        temp = ctk.CTkEntry(self,width=100,font=("Century Gothic bold",16),textvariable=temp_value,fg_color="transparent")
        temp.place(x=205,y=400)

        sat = ctk.CTkEntry(self,width=100,font=("Century Gothic bold",16),textvariable=sat_value,fg_color="transparent")
        sat.place(x=520,y=280)

        dat = ctk.CTkEntry(self,width=100,font=("Century Gothic bold",16),textvariable=dat_value,fg_color="transparent")
        dat.place(x=520,y=315)

        numero_pront = ctk.CTkEntry(self,width=140,font=("Century Gothic bold",16),textvariable=pront_value,fg_color="transparent")
        numero_pront.place(x=480,y=350)

        #Combox

        nv_box = ctk.CTkComboBox(self,values=["1","0"],width=90)
        nv_box.place(x=600,y=240)

        #labels - nome,data nascimento,p.respi,psistolica,p.diastolica,freq.cardiaca,temp,nivelcosn,saturação02.

        nome = ctk.CTkLabel(self,text="Nome:",text_color=["#000","#fff"],font=("Century Gothic bold",16))
        nome.place(x=50,y=190)
        
        data_de_nascimento = ctk.CTkLabel(self,text="Data de nascimento:",text_color=["#000","#fff"],font=("Century Gothic bold",16))
        data_de_nascimento.place(x=440,y=190)

        p_diastolica = ctk.CTkLabel(self,text="Pressão Diastólica:",text_color=["#000","#fff"],font=("Century Gothic bold",16))
        p_diastolica.place(x=50,y=240)

        p_respiratoria = ctk.CTkLabel(self,text="Pressão Respiratória:",text_color=["#000","#fff"],font=("Century Gothic bold",16))
        p_respiratoria.place(x=50,y=280)

        p_sistolica = ctk.CTkLabel(self,text="Pressão Sistólica:",text_color=["#000","#fff"],font=("Century Gothic bold",16))
        p_sistolica.place(x=50,y=320)
        
        freq_cardiaca = ctk.CTkLabel(self,text="Freq.Cardíaca :",text_color=["#000","#fff"],font=("Century Gothic bold",16))
        freq_cardiaca.place(x=50,y=360)

        temp = ctk.CTkLabel(self,text="Temperatura:",text_color=["#000","#fff"],font=("Century Gothic bold",16))
        temp.place(x=50,y=400)

        nv = ctk.CTkLabel(self,text="Nível de Consciência:",text_color=["#000","#fff"],font=("Century Gothic bold",16))
        nv.place(x=440,y=240)

        saturacao = ctk.CTkLabel(self,text="Saturação:",text_color=["#000","#fff"],font=("Century Gothic bold",16))
        saturacao.place(x=440,y=280)

        data = ctk.CTkLabel(self,text="Data Atual:",text_color=["#000","#fff"],font=("Century Gothic bold",16))
        data.place(x=440,y=315)

        np = ctk.CTkLabel(self,text="N.P:",text_color=["#000","#fff"],font=("Century Gothic bold",16))
        np.place(x=440,y=350)

        #botoes

        botao_salvar_planilha = ctk.CTkButton(self,text="Salvar".upper(),fg_color="#151",command=submit,hover_color="#131")
        botao_salvar_planilha.place(x=550,y=460)

        botao_limpar_dados = ctk.CTkButton(self,text="Limpar".upper(),fg_color="#151",command=clear,hover_color="#131")
        botao_limpar_dados.place(x=400,y=460)
        

    def change_appearance(self, new_appearance_mode):
        ctk.set_appearance_mode(new_appearance_mode)

if __name__ == "__main__":
    app = App()
    app.mainloop()
